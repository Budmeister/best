import os
import shutil
import argparse
from datetime import datetime
import re

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from antlr4 import *
from parser.BesLexer import BesLexer
from parser.BesParser import BesParser

import versioned_formulae as vf

errors = 0

def compile_formula_id_regex(ids):
    patterns = [rf"(?<![a-zA-Z_\\])\b{re.escape(id)}\b(?![a-zA-Z0-9_\\.])" for id in ids]
    combined_patern = re.compile('|'.join(patterns))
    return combined_patern

def compile_xlfn_regex():
    return compile_formula_id_regex(vf.versioned_formulae)

RED = "\033[31m"
YELLOW = "\033[33m"
RESET = "\033[0m"

def error(msg):
    global errors
    errors += 1
    print(f"{RED}ERROR: {msg}{RESET}")

def warning(msg):
    print(f"{YELLOW}WARNING: {msg}{RESET}")

def unexpected_child(child):
    if hasattr(child, "getText"):
        error(f"Unexpected {type(child)} child: {child.getText()}")
    else:
        error(f"Unexpected {type(child)} child: {child}")

a1_pattern = re.compile(r"^[a-zA-Z]{1,3}[1-9][0-9]*$")
r1c1_pattern = re.compile(r"^R[1-9][0-9]*C[1-9][0-9]*$")

def validate_name(name: str, line):
    if re.match(a1_pattern, name):
        error(f"The name, `{name}`, (line {line}) is not valid since it is an A1-style reference to a cell.")
        return
    if re.match(r1c1_pattern, name):
        error(f"The name, `{name}`, (line {line}) is not valid since it is an R1C1-style reference to a cell.")
        return
    if len(name) > 250:
        error(f"The name, `{name}`, (line {line}) is not valid since it contains more than 250 characters.")
        return

def parse_file(filename):
    with open(filename, "r") as file:
        contents = file.read()

    input_stream = InputStream(contents)

    lexer = BesLexer(input_stream)
    token_stream = CommonTokenStream(lexer)

    parser = BesParser(token_stream)
    tree = parser.file_()

    return tree

def get_file_elements_rec(filepath, imported_files=None):
    if imported_files is None:
        imported_files = set()

    if filepath in imported_files:
        return [], [], []

    parsed_file = parse_file(filepath)
    children = parsed_file.getChildren()

    import_decls = []
    expr_stms = []
    let_stms = []
    fn_stms = []

    for child in children:
        if isinstance(child, BesParser.ImportDeclContext):
            import_decls.append(child)
        elif isinstance(child, BesParser.StatementContext):
            subchildren = list(child.getChildren())
            if len(subchildren) != 1:
                raise ValueError(f"Statement object had {len(subchildren)} subchildren")
            subchild = subchildren[0]

            if isinstance(subchild, BesParser.LetStmContext):
                let_stms.append(subchild)
            elif isinstance(subchild, BesParser.ExprStmContext):
                expr_stms.append(subchild)
            elif isinstance(subchild, BesParser.FunctionStmContext):
                fn_stms.append(subchild)
            else:
                unexpected_child(subchild)
        elif isinstance(child, tree.Tree.TerminalNodeImpl) and child.getText() == "<EOF>":
            pass
        else:
            unexpected_child(child)

    current_dir = os.path.dirname(filepath)
    for import_decl in import_decls:
        import_decl: BesParser.ImportDeclContext
        identifier = import_decl.IDENTIFIER().getText()
        if identifier.startswith("\\"):
            raise ValueError(f"Illegal import name: {identifier}")
        new_filepath = os.path.join(current_dir, f"{identifier}.bes")

        new_expr_stms, new_let_stms, new_fn_stms = get_file_elements_rec(new_filepath)
        expr_stms += new_expr_stms
        let_stms += new_let_stms
        fn_stms += new_fn_stms
        
    return expr_stms, let_stms, fn_stms

def stm_to_let(stm, lets, defines, local_defines):
    if isinstance(stm, BesParser.LetStmContext):
        identifier = stm.IDENTIFIER().getText()
        expr = stm.expression()
        formula = expr_to_formula(expr, defines, local_defines)
        if identifier in lets:
            error(f"Redefinition of name `{identifier}` on line {stm.IDENTIFIER().symbol.line}")
        validate_name(identifier, stm.IDENTIFIER().symbol.line)
        lets[identifier] = formula
    elif isinstance(stm, BesParser.ExprStmContext):
        identifier = stm.IDENTIFIER().getText()
        expr = stm.expression()
        formula = expr_to_formula(expr, defines, local_defines)
        if identifier in local_defines:
            error(f"Redefinition of name \"{identifier}\" on line {stm.IDENTIFIER().symbol.line}")
        validate_name(identifier, stm.IDENTIFIER().symbol.line)
        local_defines[identifier] = formula
    elif isinstance(stm, BesParser.FunctionStmContext):
        identifier = stm.IDENTIFIER().getText()
        matchable_args = []
        args = ""
        for bracketedId in stm.idList().possiblyBracketedIdentifier():
            id = bracketedId.IDENTIFIER().getText()
            validate_name(id, bracketedId.IDENTIFIER().symbol.line)
            bracketed = bracketedId.getText().startswith("[")
            matchable_args.append(id)
            if bracketed:
                args = f"{args}_xlop.{id},"
            else:
                args = f"{args}_xlpm.{id},"
        defined_arg_regex = compile_formula_id_regex(matchable_args)
        expr = stm.blockExpr()
        body_formula = expr_to_formula(expr, defines, local_defines)
        if args:
            # Replace parameter "a" with "_xlpm.a". In Excel it still looks like "a", but in the code they store it differently.
            body_formula = defined_arg_regex.sub(lambda match: f"_xlpm.{match.group(0)}", body_formula)
        formula = f"LAMBDA({args}{body_formula})"
        if identifier in lets:
            error(f"Redefinition of name `{identifier}` on line {stm.IDENTIFIER().symbol.line}")
        validate_name(identifier, stm.IDENTIFIER().symbol.line)
        lets[identifier] = formula
    elif isinstance(stm, BesParser.StatementContext):
        let_stm = stm.letStm()
        if let_stm is not None:
            return stm_to_let(let_stm, lets, defines, local_defines)
        expr_stm = stm.exprStm()
        if expr_stm is not None:
            return stm_to_let(expr_stm, lets, defines, local_defines)
        function_stm = stm.functionStm()
        return stm_to_let(function_stm, lets, defines, local_defines)
    else:
        unexpected_child(stm)

def flatten_if_expr(if_expr):
    ifs = []
    condition = if_expr.expression()
    value_if_true = if_expr.blockExpr(0)
    ifs.append((condition, value_if_true))

    value_if_false = if_expr.blockExpr(1)
    if value_if_false is not None:
        ifs.append(("TRUE", value_if_false))
    else:
        value_if_false = if_expr.ifExpr()
        subsequent_ifs = flatten_if_expr(value_if_false)
        ifs += subsequent_ifs
    
    return ifs

def expand_definitions(string, defines, local_defines):
    while "`" in string:
        start = string.index("`")
        end = string.index("`", start+1)

        name = string[start+1:end]
        if name in local_defines:
            string = string[:start] + "(" + local_defines[name] + ")" + string[end+1:]
        elif name in defines:
            string = string[:start] + "(" + defines[name] + ")" + string[end+1:]
        else:
            error(f"Unrecognized reference to name, `{name}`, in string, \"{string}\"")
    return string

def expr_to_formula(expr, defines, local_defines=None):
    if local_defines is None:
        local_defines = {}
    if isinstance(expr, BesParser.BlockExprContext):
        local_defines = local_defines.copy()
        lets = {}

        for stm in expr.statement():
            stm_to_let(stm, lets, local_defines, defines)
        
        final_expr = expr_to_formula(expr.expression(), defines, local_defines)
        if lets:
            names_so_far = []
            formula = "LET("
            for name in lets:
                if names_so_far:
                    # Replace parameter "a" with "_xlpm.a". In Excel it still looks like "a", but in the code they store it differently.
                    defined_let_regex = compile_formula_id_regex(names_so_far)
                    value = defined_let_regex.sub(lambda match: f"_xlpm.{match.group(0)}", lets[name])
                else:
                    value = lets[name]
                formula = f"{formula}_xlpm.{name},{value},"
                names_so_far.append(name)
            defined_let_regex = compile_formula_id_regex(names_so_far)
            final_value = defined_let_regex.sub(lambda match: f"_xlpm.{match.group(0)}", final_expr)
            formula = f"{formula}{final_value})"
        else:
            formula = final_expr
        return formula
        
    elif isinstance(expr, BesParser.IfExprContext):
        ifs = flatten_if_expr(expr)
        if len(ifs) == 2:
            # Regular if
            condition, value_if_true = ifs[0]
            value_if_false = ifs[1][1]
            condition = expr_to_formula(condition, defines, local_defines)
            value_if_true = expr_to_formula(value_if_true, defines, local_defines)
            value_if_false = expr_to_formula(value_if_false, defines, local_defines)
            formula = f"IF({condition}, {value_if_true}, {value_if_false})"
        else:
            formula = "IFS("
            for condition, value_if_true in ifs:
                if condition != "TRUE":
                    condition = expr_to_formula(condition, defines, local_defines)
                value_if_true = expr_to_formula(value_if_true, defines, local_defines)
                formula = f"{formula}{condition},{value_if_true},"

            formula = f"{formula[:-1]})"
        return formula
        
    elif isinstance(expr, BesParser.IflExprContext):
        ifs = flatten_if_expr(expr)
        if len(ifs) == 2:
            # Regular if
            condition, value_if_true = ifs[0]
            value_if_false = ifs[1][1]
            condition = expr_to_formula(condition, defines, local_defines)
            value_if_true = expr_to_formula(value_if_true, defines, local_defines)
            value_if_false = expr_to_formula(value_if_false, defines, local_defines)
            formula = f"IF({condition}, LAMBDA({value_if_true}), LAMBDA({value_if_false}))()"
        else:
            formula = "IFS("
            for condition, value_if_true in ifs:
                if condition != "TRUE":
                    condition = expr_to_formula(condition, defines, local_defines)
                value_if_true = expr_to_formula(value_if_true, defines, local_defines)
                formula = f"{formula}{condition},LAMBDA({value_if_true}),"

            formula = f"{formula[:-1]})()"
        return formula

    elif isinstance(expr, tree.Tree.TerminalNodeImpl):
        if expr.getSymbol().type == BesLexer.FORMULA_LITERAL:
            formula = expr.getText()[1:-1]
            formula = bytes(formula, 'utf-8').decode('unicode_escape')
            formula = expand_definitions(formula, defines, local_defines)
        elif expr.getSymbol().type == BesLexer.STRING_LITERAL:
            formula = expr.getText()[1:]
        elif expr.getSymbol().type == BesLexer.DEFINED_EXPRESSION:
            formula = expr.getText()
            formula = expand_definitions(formula, defines, local_defines)
        elif expr.getSymbol().type == BesLexer.IDENTIFIER:
            name = expr.getText()
            if name in local_defines:
                formula = local_defines[name]
            elif name in defines:
                formula = defines[name]
            else:
                formula = name
        else:
            unexpected_child(expr)
        return formula

    elif isinstance(expr, BesParser.ExpressionContext):
        block_expr = expr.blockExpr()
        if block_expr is not None:
            return expr_to_formula(block_expr, defines, local_defines)
        if_expr = expr.ifExpr()
        if if_expr is not None:
            return expr_to_formula(if_expr, defines, local_defines)
        ifl_expr = expr.iflExpr()
        if ifl_expr is not None:
            return expr_to_formula(ifl_expr, defines, local_defines)
        formula_literal = expr.FORMULA_LITERAL()
        if formula_literal is not None:
            return expr_to_formula(formula_literal, defines, local_defines)
        string_literal = expr.STRING_LITERAL()
        if string_literal is not None:
            return expr_to_formula(string_literal, defines, local_defines)
        defined_expression = expr.DEFINED_EXPRESSION()
        if defined_expression is not None:
            return expr_to_formula(defined_expression, defines, local_defines)
        identifier = expr.IDENTIFIER()
        if identifier is not None:
            return expr_to_formula(identifier, defines, local_defines)
        expression = expr.expression()
        return expr_to_formula(expression, defines, local_defines)
    else:
        unexpected_child(expr)

def compile_file(filepath):
    global errors
    errors = 0
    expr_stms, let_stms, fn_stms = get_file_elements_rec(filepath)
    lets = {}
    defines = {}
    for expr_stm in expr_stms:
        stm_to_let(expr_stm, lets, defines.copy(), defines)
    for let_stm in let_stms:
        stm_to_let(let_stm, lets, defines.copy(), defines)
    for fn_stm in fn_stms:
        stm_to_let(fn_stm, lets, defines.copy(), defines)

    if errors != 0:
        print(f"Unable to compile due to {errors} errors")
        return None
    
    # Functions introduced into Excel after 2007 are stored in the code with the 
    # prefix, "_xlfn.", so we have to put it into the code with that prefix.
    versioned_formula_regex = compile_xlfn_regex()

    for name in lets:
        defn = lets[name]
        defn = versioned_formula_regex.sub(lambda match: f"_xlfn.{match.group(0)}", defn)
        lets[name] = defn

    return lets

BEST_MARKER = "===Compiled with Best==="

def store_lets(lets, wb, no_clear, overwrite):
    old_comments = {}
    if not no_clear:
        for name in list(wb.defined_names):
            defn = wb.defined_names[name]
            if defn.comment is not None and BEST_MARKER in defn.comment:
                old_comments[name] = defn.comment
                del wb.defined_names[name]
    
    for name in lets:
        if name in wb.defined_names and not overwrite:
            error(f"Name {name} already defined in the workbook and `overwrite` was not passed in.")
            continue
        comment = old_comments.get(name, BEST_MARKER)
        defn = DefinedName(name, comment=comment, attr_text=lets[name])
        wb.defined_names[name] = defn

def backup_file(filepath, backup_dir):
    os.makedirs(backup_dir, exist_ok=True)
    filename = os.path.basename(filepath)
    curtime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    new_filename = f"{filename[:-len('.xlsx')]}_{curtime}.xlsx"

    shutil.copyfile(filepath, os.path.join(backup_dir, new_filename))

def do(args, output_file):
    if args.do == "clear-bes-defs":
        if not args.input:
            print("ERROR: To perform this action, you must provide an input")
            return
        if not args.no_backup:
            backup_file(args.input, args.backup_dir)
        wb = load_workbook(args.input)
        store_lets({}, wb, False, False)
        wb.save(output_file)
    elif args.do == "clear-defs":
        if not args.input:
            print("ERROR: To perform this action, you must provide an input")
            return
        if not args.no_backup:
            backup_file(args.input, args.backup_dir)
        wb = load_workbook(args.input)
        for name in list(wb.defined_names):
            del wb.defined_names[name]
        wb.save(output_file)
    elif args.do == "print-defs" or args.do == "print-defs-full":
        if not args.input:
            print("ERROR: To perform this action, you must provide an input")
            return
        if not args.no_backup:
            backup_file(args.input, args.backup_dir)
        wb = load_workbook(args.input)
        for name in wb.defined_names:
            print(f"{name}: ")
            if args.do == "print-defs":
                if wb.defined_names[name].comment:
                    print(f"\tComment: {wb.defined_names[name].comment}")
                print(f"\tValue: {wb.defined_names[name].attr_text}")
            else:
                print(f"\t{wb.defined_names[name]}")
    elif args.do == "delete-backups":
        shutil.rmtree(args.backup_dir)
    else:
        error(f"Unrecognized action: {args.do}")

def main(args):
    # Validate the output file
    if args.output:
        output_file = args.output
        if not output_file.endswith(".xlsx"):
            print(f"Invalid output file \"{args.input}\" - must end with \".xlsx\"")
    elif args.input:
        output_file = args.input
    else:
        output_file = "BesBook.xlsx"

    # Validate the input file    
    if args.input:
        if not args.input.endswith(".xlsx"):
            print(f"Invalid input file \"{args.input}\" - must end with \".xlsx\"")
            return

    if args.do:
        do(args, output_file)
        return
    
    if not args.script:
        print("ERROR: Expected either --script or --do to be specified")
        return

    lets = compile_file(args.script)
    if lets is None:
        return

    if args.input:
        if not args.no_backup:
            backup_file(args.input, args.backup_dir)
        wb = load_workbook(args.input)
    else:
        wb = Workbook()

    store_lets(lets, wb, args.no_clear, args.overwrite_defs)

    if errors != 0:
        print(f"Unable to save the file because of {errors} errors")
        return
    
    wb.save(output_file)
    

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Transpiler from Bud Excel Script to Excel Workbook.")
    
    parser.add_argument("-s", "--script", help="The name of the Bes file to compile")
    
    # Config
    parser.add_argument("-i", "--input", help="The Excel file to modify")
    parser.add_argument("-o", "--output", help="The output Excel file (defaults to input file if specified, or 'BesBook.xlsx')")
    parser.add_argument("-b", "--backup-dir", dest="backup_dir", help="The directory to put backups (defaults to ./backups)", default="./backups")
    parser.add_argument("--no-backup", dest="no_backup", help="Do not backup the input file before overwriting it", action="store_true")
    parser.add_argument("--no-clear-defs", dest="no_clear", help="Do not remove old definitions created by Best (non-Best definitions are unaffected)", action="store_true")
    parser.add_argument("--overwrite-defs", dest="overwrite_defs", help="Overwrite existing definitions", action="store_true")
    parser.add_argument("-d", "--do", help="Do an action instead of compiling a script", choices=["clear-bes-defs", "clear-defs", "print-defs", "print-defs-full", "delete-backups"])
    
    args = parser.parse_args()
    main(args)

